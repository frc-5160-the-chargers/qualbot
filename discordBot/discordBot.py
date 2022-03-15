import json, discord, requests, random, math, openpyxl
from discord.ext import commands
from openpyxl import Workbook

with open(r"discordBot\auth.json") as auth_file:
    authFile = json.load(auth_file)
    TOKEN = authFile["token"]
    authKey = authFile["API"]
    guildID = authFile["ID"]

with open(r"discordBot\scouters.txt", "r") as scouterList:
    scoutingPairs = scouterList.read().split("\n")

bot = commands.Bot(command_prefix='!', description="Bot commands")

@bot.command(name="createChannels", description="Create channels")
async def createChannels(ctx, event):
    teamList = getTeams(event)
    for i in range(len(teamList)):
        teamInfo = requests.get(f"https://www.thebluealliance.com/api/v3/team/frc{teamList[i]}/simple", headers={"X-TBA-Auth-Key": authKey}).json()
        teamList[i] += "_" + teamInfo["nickname"] + f"_{event}"

    await ctx.guild.create_category(event)
    category = discord.utils.get(ctx.guild.categories, name=event)
    for i in range(len(teamList)):
        #These lines work, I would not recommend running them until they are needed
        await ctx.guild.create_text_channel(teamList[i], category=category)

    teams = randomizeTeams(teamList, len(scoutingPairs))
    scoutsTeams = {}
    for i in range(len(scoutingPairs)):
        scoutsTeams.setdefault(scoutingPairs[i], teams[i])
    with open(r"discordBot\teamsRandomized.json", "w") as writeFile:
        writeFile.write(str(scoutsTeams))

@bot.command(name="deleteChannels", description="Create channels")
async def deleteChannels(ctx, event):
    # teamList = getTeams(event)
    # for i in range(len(teamList)):
    #     teamInfo = requests.get(f"https://www.thebluealliance.com/api/v3/team/frc{teamList[i]}/simple", headers={"X-TBA-Auth-Key": authKey}).json()
    #     teamList[i] += "_" + teamInfo["nickname"] + f"_{event}"

    # text_channels = []
    # for channel in ctx.guild.text_channels:
    #     text_channels.append(channel)
    
    # deletion_channels = []
    # for channel in text_channels:
    #     if channel[]
    # await ctx.guild.create_category(event)
    # category = discord.utils.get(ctx.guild.categories, name=event)
    # for i in range(len(teamList)):
        #These lines work, I would not recommend running them until they are needed
        # channel = "#2059_the-hitchhikers_2020ncwak"
        # await channel.delete()
    pass

def getTeams(event):
    teams = requests.get(url=f"https://www.thebluealliance.com/api/v3/event/{event}/teams", headers={"X-TBA-Auth-Key": authKey}).json()
    teamList = []
    for team in teams:
        teamList.append(team["key"][3:])
    teamList.remove("5160")
    return teamList

def randomizeTeams(teamList, scoutpairs=6):
    random.shuffle(teamList)
    teamRandom = []
    teamsPerPair = math.ceil(len(teamList) / scoutpairs)
    for i in range(0, len(teamList), teamsPerPair):
        teamRandom.append(teamList[i:i + teamsPerPair])
    if len(teamRandom[-1]) != teamsPerPair:
        randomChoice = random.choice(teamRandom[-2])
        teamRandom[-2].remove(randomChoice)
        teamRandom[-1].append(randomChoice)
    return teamRandom

@bot.command(name='checkScouting', description="Reads through channels under a specific event ")
async def checkScouting(ctx, event):
    info = ["Climbing", 'Shooter Type', 'Automonus', "Role", "Climbing Time", "Other Notes"]
    eventName = input("Enter event name: ")
    server = bot.get_guild(guildID)
    wb = Workbook()
    for channel in server:
        if channel.name == eventName:
            rowNum = 2
            colNum = 2
            sheet = wb.create_sheet(title=channel.name)
            for textChannel in range(len(channel.channels)):
                sheet.cell(column=1, row=rowNum, value=textChannel)   
                rowNum+=1
            rowNum = 2
            for role in info:
                sheet.cell(column=colNum, row=1, value=role)
                colNum+=1

bot.run(TOKEN)